"""Export dpo/runs to Excel grouped by config_snapshot settings (one sheet per setting)."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
RUNS_DIR = ROOT / "dpo" / "runs"

import sys

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from dpo.types import is_episode_trajectory_file, trajectories_from_episode_file


def _load_json(path: Path) -> dict | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _load_config(run_dir: Path) -> dict:
    snap = run_dir / "config_snapshot.yaml"
    if snap.exists():
        return yaml.safe_load(snap.read_text(encoding="utf-8")) or {}
    info = _load_json(run_dir / "run_info.json") or {}
    return info.get("resolved_config") or info.get("config") or {}


def _discover_run_dirs() -> list[Path]:
    """Leaf run folders that contain trajectories."""
    found: set[Path] = set()
    if not RUNS_DIR.is_dir():
        return []
    for traj in RUNS_DIR.rglob("trajectories"):
        if not traj.is_dir():
            continue
        if not any(traj.glob("*.json")):
            continue
        found.add(traj.parent.resolve())
    return sorted(found, key=lambda p: str(p.relative_to(RUNS_DIR)))


def _run_label(run_dir: Path) -> str:
    try:
        return str(run_dir.relative_to(RUNS_DIR)).replace("\\", "/")
    except ValueError:
        return run_dir.name


def _model_tag(model: str) -> str:
    if not model:
        return "?"
    return model.split(":")[-1].replace(".", "")


def _config_fingerprint(cfg: dict) -> str:
    agents = cfg.get("agents") or {}
    blind = agents.get("blind") or {}
    oracle = agents.get("oracle") or {}
    payload = {
        "env": cfg.get("env", ""),
        "horizon": (cfg.get("env_params") or {}).get("horizon"),
        "n_episodes": cfg.get("n_episodes"),
        "blind_model": blind.get("model"),
        "blind_temp": blind.get("temperature"),
        "oracle_model": oracle.get("model"),
        "oracle_temp": oracle.get("temperature"),
        "pairing_mode": cfg.get("pairing_mode"),
        "min_payoff_gap": cfg.get("min_payoff_gap"),
        "min_alignment_gap": cfg.get("min_alignment_gap"),
        "role_tolerance": cfg.get("role_tolerance"),
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(raw.encode()).hexdigest()[:12]


def _setting_sheet_name(cfg: dict) -> str:
    env = (cfg.get("env") or "env").replace("-", "")
    agents = cfg.get("agents") or {}
    blind = agents.get("blind") or {}
    oracle = agents.get("oracle") or {}
    b_m = _model_tag(str(blind.get("model", "")))
    o_m = _model_tag(str(oracle.get("model", "")))
    b_t = blind.get("temperature", "")
    o_t = oracle.get("temperature", "")
    h = (cfg.get("env_params") or {}).get("horizon", "")
    name = f"{env}_b{b_m}_{b_t}_o{o_m}_{o_t}_h{h}"
    name = re.sub(r"[^\w\-.]", "_", name)
    return name[:31]


def _unique_sheet_names(groups: dict[str, list[Path]], cfg_by_fp: dict[str, dict]) -> dict[str, str]:
    """Map fingerprint -> unique Excel sheet name."""
    used: dict[str, int] = {}
    out: dict[str, str] = {}
    for fp in groups:
        base = _setting_sheet_name(cfg_by_fp[fp])
        n = used.get(base, 0)
        used[base] = n + 1
        sheet = base if n == 0 else f"{base[:28]}_{n + 1}"
        out[fp] = sheet[:31]
    return out


def _episode_rows(run_dir: Path, cfg: dict) -> list[dict]:
    traj_dir = run_dir / "trajectories"
    if not traj_dir.is_dir():
        return []

    info = _load_json(run_dir / "run_info.json") or {}
    if cfg.get("n_episodes") is None:
        cfg = {**cfg, "n_episodes": (info.get("resolved_config") or {}).get("n_episodes")}

    agents = cfg.get("agents") or {}
    blind_model = (agents.get("blind") or {}).get("model", "")
    oracle_model = (agents.get("oracle") or {}).get("model", "")
    run_path = _run_label(run_dir)

    seen: set[tuple[str, str]] = set()
    rows: list[dict] = []

    traj_paths = sorted(traj_dir.glob("ep*.json"))
    for path in traj_paths:
        raw = json.loads(path.read_text(encoding="utf-8"))
        if is_episode_trajectory_file(raw):
            m = re.match(r"ep(\d+)_(blind|oracle)\.json", path.name)
            if not m:
                continue
            ep_idx, role = m.group(1), m.group(2)
            key = (ep_idx, role)
            if key in seen:
                continue
            seen.add(key)
            trajs = {t.player: t for t in trajectories_from_episode_file(raw)}
            p0 = trajs.get("p0")
            p1 = trajs.get("p1")
            if p0 is None:
                continue
            data = json.loads(p0.to_json())
            meta = data.get("meta") or {}
            run_meta = meta.get("run_meta") or meta
            opponent = run_meta.get("opponent_id") or meta.get("opponent_id", "")
            actions = "".join(a.get("value", "?") for a in data.get("actions") or [])
            decision_meta = meta.get("decision_meta") or []
            malformed = sum(1 for d in decision_meta if d.get("malformed"))
            reward_p1 = p1.cumulative_reward if p1 is not None else None
        else:
            m = re.match(r"ep(\d+)_(blind|oracle)_p0\.json", path.name)
            if not m:
                continue
            ep_idx, role = m.group(1), m.group(2)
            key = (ep_idx, role)
            if key in seen:
                continue
            seen.add(key)
            data = raw
            meta = data.get("meta") or {}
            run_meta = meta.get("run_meta") or meta
            opponent = run_meta.get("opponent_id") or meta.get("opponent_id", "")
            actions = "".join(a.get("value", "?") for a in data.get("actions") or [])
            decision_meta = meta.get("decision_meta") or []
            malformed = sum(1 for d in decision_meta if d.get("malformed"))
            p1_path = path.with_name(path.name.replace("_p0.json", "_p1.json"))
            reward_p1 = None
            if p1_path.exists():
                p1 = json.loads(p1_path.read_text(encoding="utf-8"))
                reward_p1 = p1.get("cumulative_reward")

        rows.append(
            {
                "run_path": run_path,
                "episode": int(ep_idx),
                "role": role,
                "opponent": opponent,
                "env": cfg.get("env") or data.get("env_id", ""),
                "horizon": (cfg.get("env_params") or {}).get("horizon"),
                "n_episodes": cfg.get("n_episodes"),
                "reward_p0": data.get("cumulative_reward"),
                "reward_p1": reward_p1,
                "total_reward": (
                    (data.get("cumulative_reward") or 0) + (reward_p1 or 0)
                    if reward_p1 is not None
                    else None
                ),
                "actions_p0": actions,
                "rounds": len(data.get("actions") or []),
                "malformed_steps": malformed,
                "blind_model": blind_model,
                "oracle_model": oracle_model,
                "blind_temp": (agents.get("blind") or {}).get("temperature"),
                "oracle_temp": (agents.get("oracle") or {}).get("temperature"),
                "run_id": data.get("run_id", ""),
            }
        )
    return rows


def _pair_rows(run_dir: Path) -> list[dict]:
    rows: list[dict] = []
    run_path = _run_label(run_dir)
    for pairs_path in sorted(run_dir.glob("pairs_*.jsonl")):
        strategy = pairs_path.stem.replace("pairs_", "")
        for line_no, line in enumerate(pairs_path.read_text(encoding="utf-8").splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            meta = rec.get("meta") or {}
            rows.append(
                {
                    "run_path": run_path,
                    "strategy": strategy,
                    "pair_no": line_no,
                    "env": meta.get("env_id", ""),
                    "chosen_role": meta.get("chosen_role", ""),
                    "rejected_role": meta.get("rejected_role", ""),
                    "reward_chosen": meta.get("cumulative_reward_chosen"),
                    "reward_rejected": meta.get("cumulative_reward_rejected"),
                    "chosen_alignment": meta.get("chosen_alignment"),
                    "rejected_alignment": meta.get("rejected_alignment"),
                    "alignment_gap": meta.get("gap"),
                    "reason": meta.get("reason", ""),
                }
            )
    return rows


def _setting_index_rows(
    fp: str,
    sheet_name: str,
    cfg: dict,
    run_dirs: list[Path],
) -> dict:
    agents = cfg.get("agents") or {}
    blind = agents.get("blind") or {}
    oracle = agents.get("oracle") or {}
    return {
        "sheet": sheet_name,
        "fingerprint": fp,
        "env": cfg.get("env", ""),
        "horizon": (cfg.get("env_params") or {}).get("horizon"),
        "n_episodes": cfg.get("n_episodes"),
        "blind_model": blind.get("model"),
        "blind_temp": blind.get("temperature"),
        "oracle_model": oracle.get("model"),
        "oracle_temp": oracle.get("temperature"),
        "pairing_mode": cfg.get("pairing_mode"),
        "run_count": len(run_dirs),
        "run_paths": "; ".join(_run_label(d) for d in run_dirs),
    }


def _style_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if not df.empty:
        ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 48))


def export(out_path: Path | None = None) -> Path:
    run_dirs = _discover_run_dirs()
    if not run_dirs:
        raise SystemExit(f"No run folders with trajectories under {RUNS_DIR}")

    cfg_by_dir: dict[Path, dict] = {}
    groups: dict[str, list[Path]] = {}
    cfg_by_fp: dict[str, dict] = {}

    for run_dir in run_dirs:
        cfg = _load_config(run_dir)
        info = _load_json(run_dir / "run_info.json") or {}
        resolved = info.get("resolved_config") or {}
        if cfg.get("n_episodes") is None and resolved.get("n_episodes") is not None:
            cfg = {**cfg, "n_episodes": resolved["n_episodes"]}
        cfg_by_dir[run_dir] = cfg
        fp = _config_fingerprint(cfg)
        cfg_by_fp.setdefault(fp, cfg)
        groups.setdefault(fp, []).append(run_dir)

    sheet_names = _unique_sheet_names(groups, cfg_by_fp)

    index_rows = [
        _setting_index_rows(fp, sheet_names[fp], cfg_by_fp[fp], dirs)
        for fp, dirs in sorted(groups.items(), key=lambda x: sheet_names[x[0]])
    ]

    out_path = out_path or (RUNS_DIR / f"dpo_runs_by_setting_{datetime.now():%Y%m%d}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_index = pd.DataFrame(index_rows)
        df_index.to_excel(writer, sheet_name="설정_목록", index=False)
        _style_sheet(writer.sheets["설정_목록"], df_index)

        for fp, dirs in sorted(groups.items(), key=lambda x: sheet_names[x[0]]):
            ep_rows: list[dict] = []
            pair_rows: list[dict] = []
            for run_dir in sorted(dirs, key=_run_label):
                cfg = cfg_by_dir[run_dir]
                ep_rows.extend(_episode_rows(run_dir, cfg))
                pair_rows.extend(_pair_rows(run_dir))

            sheet = sheet_names[fp]
            df_ep = pd.DataFrame(ep_rows)
            if df_ep.empty:
                df_ep = pd.DataFrame([{"note": "no episodes", "run_paths": "; ".join(_run_label(d) for d in dirs)}])
            df_ep.to_excel(writer, sheet_name=sheet, index=False)
            _style_sheet(writer.sheets[sheet], df_ep)

            if pair_rows:
                pair_sheet = f"{sheet[:25]}_pairs"[:31]
                used = {sheet, "설정_목록"}
                n = 2
                while pair_sheet in used:
                    pair_sheet = f"{sheet[:22]}_pairs{n}"[:31]
                    n += 1
                df_pairs = pd.DataFrame(pair_rows)
                df_pairs.to_excel(writer, sheet_name=pair_sheet, index=False)
                _style_sheet(writer.sheets[pair_sheet], df_pairs)
                used.add(pair_sheet)

    return out_path


if __name__ == "__main__":
    path = export()
    print(f"Wrote {path}")
